import openpyxl
import sqlite3


def read_table(table_name, file_name, date):
    conn = sqlite3.connect(f'databases/{table_name}.db', check_same_thread=False, timeout=30)
    wb = openpyxl.load_workbook(f'protocols/{file_name}')
    ws = wb.active
    rows = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
    my_list = list()
    new_list = list()
    list_of_tables = list()
    for value in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column,
            values_only=True):
        my_list.append(value)
    for i in range(0, len(rows)):
        if str(rows[i]).startswith('М') or str(rows[i]).startswith('Д') or str(rows[i]).startswith('Ю') or str(rows[i]).startswith('Ж'):
            new_list.append(i + 2)
            list_of_tables.append(str(rows[i]))
            create_table(str(rows[i]), conn)
    new_list.append(ws.max_row)
    conn.cursor().execute("""CREATE TABLE IF NOT EXISTS info (
           id INTEGER NOT NULL,
           date TEXT NOT NULL,
           status INTEGER NOT NULL,
           PRIMARY KEY(id));""")
    conn.commit()
    conn.cursor().execute("INSERT INTO info (id, date, status) VALUES(?, ?, ?)", (0, date, 0))
    conn.commit()
    name = ''
    for row in my_list:
        if str(row[0]).startswith('Д') or str(row[0]).startswith('С') or str(row[0]).startswith('Ю') or str(
                row[0]).startswith('М') or str(row[0]).startswith('Ж') or row[0] is None:
            for elem in list_of_tables:
                if elem == str(row[0]):
                    name = elem
        else:
            query = (
                "INSERT INTO {name} (id, startNumber, fio, dateBirth, team, startTime, finishTime, result, gap, "
                "place, distance) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)").format(name=name.replace(' ',
                                                                                        '').split(
                'г')[0].replace('-', ''))
            conn.cursor().execute(query, (
                row[0], row[0], row[1], row[2], row[3],
                str(row[4]), row[5], None, None, None, name))
            conn.commit()

    # param = '00:02:00'
    # conn.cursor().execute("UPDATE Девочки20092010 SET result = ? WHERE id = ?", (param, "36"))
    # conn.cursor().execute("""CREATE TABLE temp_table (
    #     id INTEGER PRIMARY KEY,
    #     startNumber INTEGER NOT NULL,
    #     fio TEXT NOT NULL,
    #     dateBirth TEXT NOT NULL,
    #     team TEXT NOT NULL,
    #     startTime TEXT NOT NULL,
    #     finishTime TEXT,
    #     result TEXT,
    #     gap TEXT,
    #     place INTEGER)""")
    # conn.cursor().execute("""INSERT INTO temp_table (startNumber, fio, dateBirth, team, startTime, finishTime,
    # result, gap, place) SELECT startNumber, fio, dateBirth, team, startTime, finishTime, result, gap, place FROM
    # Девочки20092010 ORDER BY case when result is null then 1 else 0 end, result""")
    # conn.cursor().execute("""DROP TABLE Девочки20092010""")
    # conn.cursor().execute("""ALTER TABLE temp_table RENAME TO Девочки20092010""")
    # conn.commit()
    conn.close()


def create_table(table_name, conn):
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS {tab} (
   id INTEGER NOT NULL,
   startNumber INTEGER NOT NULL,
   fio TEXT NOT NULL,
   dateBirth TEXT NOT NULL,
   team TEXT NOT NULL,
   startTime TEXT NOT NULL,
   finishTime TEXT,
   result TEXT,
   gap TEXT,
   place INTEGER,
   distance TEXT NOT NULL,
   PRIMARY KEY(id));""".format(tab=table_name.replace(' ', '').split('г')[0]).replace('-', ''))
    conn.commit()


class Sportsman:
    def __init__(self, startNumber, fio, dateBirth, team, startTime, finishTime, result, gap, place):
        self.startNumber = startNumber
        self.fio = fio
        self.dateBirth = dateBirth
        self.team = team
        self.startTime = startTime
        self.finishTime = finishTime
        self.result = result
        self.gap = gap
        self.place = place
